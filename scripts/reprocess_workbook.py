from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app.dedup import deduplicate
from app.extractor import MappingResolver, to_structured
from app.matcher import compute_matches, compute_priority, demand_summary, supply_summary, top_leads
from app.schemas import (
    CLEAN_DATA_COLUMNS,
    MATCH_COLUMNS,
    STRUCTURED_COLUMNS,
    SUMMARY_DEMAND_COLUMNS,
    SUMMARY_SUPPLY_COLUMNS,
    TOP_LEAD_COLUMNS,
    ParsedMessage,
)
from app.sheets_client import DEFAULT_CONFIG, DEFAULT_WEIGHTS


REJECTED_SHEET_NAMES = ("Rejected / Ignored Data", "Rejected  Ignored Data")


def _sheet_by_name(workbook, name: str):
    if name in workbook.sheetnames:
        return workbook[name]
    raise KeyError(f"Missing sheet: {name}")


def _rejected_sheet(workbook):
    for name in REJECTED_SHEET_NAMES:
        if name in workbook.sheetnames:
            return workbook[name]
    return workbook.create_sheet(REJECTED_SHEET_NAMES[0])


def _clear_and_write(ws, header: list[str], rows: Iterable[list[object]]) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.append(header)
    for row in rows:
        ws.append(list(row))


def _read_key_values(ws) -> dict[str, float]:
    values: dict[str, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        try:
            values[str(row[0]).strip()] = float(row[1])
        except Exception:
            continue
    return values


def _read_table(ws) -> list[list[str]]:
    return [[("" if value is None else str(value)) for value in row] for row in ws.iter_rows(values_only=True)]


def _parse_raw_messages(ws) -> list[ParsedMessage]:
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    idx = {name: pos for pos, name in enumerate(header)}
    messages: list[ParsedMessage] = []
    for row in rows[1:]:
        if not row or all(value in (None, "") for value in row):
            continue
        timestamp = row[idx["Timestamp"]]
        if not isinstance(timestamp, datetime):
            timestamp = datetime.fromisoformat(str(timestamp))
        sender = "" if row[idx["Sender"]] is None else str(row[idx["Sender"]])
        message = "" if row[idx["Raw Message"]] is None else str(row[idx["Raw Message"]])
        source = "" if row[idx["Source"]] is None else str(row[idx["Source"]])
        messages.append(
            ParsedMessage(
                timestamp=timestamp,
                sender=sender,
                message=message,
                raw_message=message,
                source=source,
            )
        )
    return messages


def main() -> None:
    parser = argparse.ArgumentParser(description="Reprocess a workbook from Raw Data using the current extraction pipeline.")
    parser.add_argument("workbook", help="Path to the workbook to rebuild")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    wb = load_workbook(workbook_path)

    raw_ws = _sheet_by_name(wb, "Raw Data")
    structured_ws = _sheet_by_name(wb, "Structured Data")
    clean_ws = _sheet_by_name(wb, "Clean Data")
    matches_ws = _sheet_by_name(wb, "Matches")
    top_ws = _sheet_by_name(wb, "Top Leads")
    demand_ws = _sheet_by_name(wb, "Demand Summary")
    supply_ws = _sheet_by_name(wb, "Supply Summary")
    rejected_ws = _rejected_sheet(wb)

    config = {**DEFAULT_CONFIG, **_read_key_values(_sheet_by_name(wb, "Config"))}
    weights = {**DEFAULT_WEIGHTS, **_read_key_values(_sheet_by_name(wb, "Scoring Weights"))}
    location_map = MappingResolver(_read_table(_sheet_by_name(wb, "Location Mapping")))
    property_map = MappingResolver(_read_table(_sheet_by_name(wb, "Property Type Mapping")))

    parsed = _parse_raw_messages(raw_ws)
    incoming = to_structured(parsed, location_map, property_map, weights)
    deduped = deduplicate([], incoming, int(config.get("dedup_window_days", 1)))
    now = datetime.now()

    matches = compute_matches(deduped.leads, weights, float(config.get("match_threshold", 55)), now)
    compute_priority(deduped.leads, matches, weights, now)
    top = top_leads(deduped.leads, int(config.get("top_leads_count", 10)))
    demand = demand_summary(deduped.leads)
    supply = supply_summary(deduped.leads)

    _clear_and_write(structured_ws, STRUCTURED_COLUMNS, [lead.to_row() for lead in deduped.leads])
    _clear_and_write(rejected_ws, STRUCTURED_COLUMNS, [lead.to_row() for lead in deduped.leads if lead.values.get("Type") == "Ignore"])
    _clear_and_write(
        clean_ws,
        CLEAN_DATA_COLUMNS,
        [[lead.values.get(column, "") for column in CLEAN_DATA_COLUMNS] for lead in deduped.leads if lead.values.get("data_status") == "APPROVED"],
    )
    _clear_and_write(matches_ws, MATCH_COLUMNS, [[row.get(column, "") for column in MATCH_COLUMNS] for row in matches])
    _clear_and_write(top_ws, TOP_LEAD_COLUMNS, [[lead.values.get(column, "") for column in TOP_LEAD_COLUMNS] for lead in top])
    _clear_and_write(demand_ws, SUMMARY_DEMAND_COLUMNS, [[row.get(column, "") for column in SUMMARY_DEMAND_COLUMNS] for row in demand])
    _clear_and_write(supply_ws, SUMMARY_SUPPLY_COLUMNS, [[row.get(column, "") for column in SUMMARY_SUPPLY_COLUMNS] for row in supply])

    wb.save(workbook_path)


if __name__ == "__main__":
    main()
